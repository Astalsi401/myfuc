import os
import logging
import re
import pandas as pd
from html import unescape
from bs4 import BeautifulSoup as bs
from base64 import b64decode, b64encode
from csv import writer, reader
from datetime import date
from json import dump
from openpyxl import load_workbook
from pikepdf import Pdf
from typing import List as TList


def cpath(p):
    if not os.path.isdir(p):
        os.makedirs(p)
    return p


def alpha(alpha):
    '''26進位英數互換'''
    if type(alpha) == str:
        alpha = alpha.upper()
        assert (isinstance(alpha, str))
        return sum([(ord(n) - 64) * 26**i for i, n in enumerate(list(alpha)[::-1])])
    elif type(alpha) == int:
        assert (isinstance(alpha, int) and alpha > 0)
        num = [chr(i) for i in range(65, 91)]
        ret = []
        while alpha > 0:
            alpha, m = divmod(alpha - 1, len(num))
            ret.append(num[m])
        return ''.join(ret[::-1])


class List:
    def __init__(self, data: list):
        self.data = data

    def writeXlsx(self, f, sheet: str, start=['A', '1']):
        '''
        引入openpyxl後再使用
        list to excel
        f = openpyxl.load_workbook(path)
        sheet = sheet name
        start = excel開始的位置
        '''
        try:
            ws = f[sheet]
        except KeyError:
            f.create_sheet(sheet, 0)
            ws = f[sheet]
        start = start[0] + start[1]
        end = alpha(alpha(start[0]) + len(self.data[0]) - 1) + str(int(start[1]) + len(self.data) - 1)
        for i, r in enumerate(ws[start:end]):
            for j, c in enumerate(r):
                c.value = self.data[i][j]

    def writeCsv(self, path: str, mode='w+', enc='utf-8-sig'):
        '''list to csv'''
        with open(f'{path}', mode=mode, encoding=enc, newline='') as f:
            for a in self.data:
                writer(f).writerow(a)

    def writeHtml(self, path: str, name: str, encoding='utf-8'):
        '''list to html table'''
        cpath(path)
        htmlTable = '<table>'
        for row in self.data:
            htmlTable += '<tr>'
            for c in row:
                htmlTable += f'<td>{c}</td>'
            htmlTable += '</tr>'
        htmlTable += '</table>'
        with open(f'{path}/{name}.html', mode='w+', encoding=encoding) as f:
            f.write(htmlTable)

    def convertToJson(self, head=False):
        '''list to json'''
        return [{self.data[0][i]: c for i, c in enumerate(a)} for a in self.data[1:]] if head else [{i: c for i, c in enumerate(a)} for a in self.data]


def xlsx(f, sheet, data, start):
    '''
    引入openpyxl後再使用
    list to excel
    f = openpyxl.load_workbook(path)
    sheet = sheet name
    data = 資料(list)
    start = excel開始的位置，如['A', '1']
    '''
    try:
        ws = f[sheet]
    except KeyError:
        f.create_sheet(sheet, 0)
        ws = f[sheet]
    start = start[0] + start[1]
    end = alpha(alpha(start[0]) + len(data[0]) - 1) + str(int(start[1]) + len(data) - 1)
    for i, r in enumerate(ws[start:end]):
        for j, c in enumerate(r):
            c.value = data[i][j]


def enc(path, name, enc):
    '''變更檔案編碼'''
    cpath(f'{path}/convert')
    with open(f'{path}/{name}', mode='r', encoding=enc[0]) as f:
        data = f.read()
    with open(f'{path}/convert/{name}', mode='w+', encoding=enc[1]) as f:
        f.write(data)


def writeCsv(path, name, data, mode='w+', enc='utf-8-sig'):
    '''list to csv'''
    cpath(path)
    with open(f'{path}/{name}', mode=mode, encoding=enc, newline='') as f:
        for a in data:
            writer(f).writerow(a)
    logging.info(f'{name} saved!')


def readCsv(path, enc='utf-8-sig'):
    '''csv to list'''
    with open(path, mode='r', encoding=enc, newline='') as f:
        return [a for a in reader(f)]


def writeJson(path, name, data, mode='w+', encoding='utf-8-sig'):
    '''write to json'''
    cpath(f'{path}/json')
    with open(f'{path}/{name}', mode, encoding=encoding) as f:
        dump(data, f, ensure_ascii=False)
    logging.info(f'{path}/{name} saved!')


def writeHtml(path, data, name, encoding='utf-8'):
    '''list to html table'''
    cpath(path)
    htmlTable = '<table>'
    for row in data:
        htmlTable += '<tr>'
        for c in row:
            htmlTable += f'<td>{c}</td>'
        htmlTable += '</tr>'
    htmlTable += '</table>'
    with open(f'{path}/{name}.html', mode='w+', encoding=encoding) as f:
        f.write(htmlTable)


def convertToJson(data):
    '''list to json'''
    json = []
    for a in data:
        i = 0
        b = {}
        for c in a:
            b[f'row{i}'] = c
            i += 1
        json.append(b)
    return json


def getXlsxSheets(path):
    '''抓取Excel分頁名稱'''
    return load_workbook(path).get_sheet_names()


def getFilesName(path, ext=None):
    '''抓取資料夾內檔案名稱, ext指定副檔名'''
    return [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f)) and f'.{ext}' in f] if ext else [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f))]


def yearsCalc(yearsAgo=0):
    '''列出最近N年年份'''
    return [str(date.today().year - a) for a in range(0, yearsAgo)]


def b64Decode(path, name, data):
    cpath(path)
    with open(f'{path}/{name}', 'wb') as f:
        f.write(b64decode(data))
        logging.info(f'{name} saved!')


def b64Encode(path, name):
    with open(f'{path}/{name}', 'rb') as f:
        return b64encode(f.read())


class MyPdf:
    def __init__(self, df: pd.DataFrame, pdfPath: str) -> None:
        self.df = df
        self.pdf = Pdf.open(pdfPath)

    def split(self, per: int, exportPath: str, fileName: TList[str], folder=None, join='-', start=0, limit=150):
        self.df['fileName'] = self.df[fileName].apply(lambda x: join.join(x.astype(str)).replace('|', ''), axis=1)
        if folder:
            self.df['folder'] = self.df[folder].str.replace('|', '', regex=False)
        for i, row in self.df.iterrows():
            if folder:
                path = f'{cpath(f"{exportPath}/{row.folder}")}/{row.fileName}.pdf' if limit == None else f'{cpath(f"{exportPath}/p{part}/{row.folder}")}/{row.fileName}.pdf'
            else:
                path = f'{cpath(exportPath)}/{row.fileName}.pdf' if limit == None else f'{cpath(f"{exportPath}/p{part}")}/{row.fileName}.pdf'
            part = None if limit == None else int(i // limit + 1)
            export = Pdf.new()
            for page in self.pdf.pages[start:start + per]:
                export.pages.append(page)
            export.save(path)
            start = start + per


class Combine:
    def __init__(self, profileDf, urlDf, sheetName: str, keep='Folder') -> None:
        self.sheet = sheetName
        self.profileDf = profileDf
        self.urlDf = urlDf[['Name', 'URL']]

    def merge(self, left_on: str, how='inner'):
        return self.profileDf.merge(self.urlDf, left_on=left_on, right_on='Name', how=how, indicator=True)


class Script:
    def __init__(self, script: str, type: str = None) -> None:
        self.script = script
        self.type = type


class HtmlContent:
    def __init__(self, html: str, css: str = None, js: TList[Script] = None) -> None:
        self.html = unescape(html)
        self.soup = bs(self.html, 'html.parser')
        self.css = f'<style>{css}</style>' if css else ''
        self.js = ''.join([f'''<script {f'type="{script.type}"' if script.type else ''}>{script.script}</script>''' for script in js]) if js else ''
        self.simp = unescape(f'{self.css}{self.body()}{self.js}')
        self.title = self.soup.title.string

    def body(self):
        return re.sub(r'<body>|</body>', '', str(self.soup.select('body')[0]))
